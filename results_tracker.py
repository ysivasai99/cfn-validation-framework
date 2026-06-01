import csv
import time
import os
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

RESULTS_DIR = Path("./research_results")
EXCEL_FILE  = RESULTS_DIR / "cfn_research_results.xlsx"
CSV_FILE    = RESULTS_DIR / "cfn_research_results.csv"

RESULTS_DIR.mkdir(exist_ok=True)

COLOR_HEADER   = "1E2761"
COLOR_CLAUDE   = "534AB7"
COLOR_GPT4     = "185FA5"
COLOR_GEMINI   = "0F6E56"
COLOR_PASS     = "EAF3DE"
COLOR_FAIL     = "FCEBEB"
COLOR_WARN     = "FAEEDA"
COLOR_ROW_ALT  = "F8F8FC"

HEADERS = [
    "Exp ID", "Timestamp", "Template Type", "Model",
    "First Pass?", "Iterations", "Errors Found",
    "Error Types", "Final Status", "Time (sec)",
    "AWS Cost Saved", "Notes"
]

def make_record(experiment_id, template_type, model, first_pass,
                iterations, errors_found, error_types, final_status,
                time_seconds, aws_cost_saved, notes=""):
    return {
        "experiment_id"  : experiment_id,
        "timestamp"      : datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "template_type"  : template_type,
        "model"          : model,
        "first_pass"     : "YES" if first_pass else "NO",
        "iterations"     : iterations,
        "errors_found"   : errors_found,
        "error_types"    : ", ".join(error_types) if error_types else "None",
        "final_status"   : final_status,
        "time_seconds"   : round(time_seconds, 2),
        "aws_cost_saved" : f"${aws_cost_saved:.4f}",
        "notes"          : notes,
    }

def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Results"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:L1")
    ws["A1"] = "AI-CFN Validation Framework — Research Results"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:L2")
    ws["A2"] = f"Masters Research 2026  |  Generated: {datetime.now().strftime('%Y-%m-%d')}"
    ws["A2"].font = Font(italic=True, size=9, color="444444")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2C3E7A")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[3].height = 22

    widths = [12, 20, 18, 10, 10, 10, 12, 25, 12, 10, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    # Summary sheet
    ws2 = wb.create_sheet("Summary Dashboard")
    ws2.merge_cells("A1:B1")
    ws2["A1"] = "Research Summary Dashboard"
    ws2["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER)
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 24

    summary = [
        ("Total Experiments",        "=COUNTA('All Results'!A4:A1000)"),
        ("First Pass Success Rate",  "=COUNTIF('All Results'!E4:E1000,\"YES\")/COUNTA('All Results'!E4:E1000)"),
        ("Avg Iterations (Claude)",  "=AVERAGEIF('All Results'!D4:D1000,\"Claude\",'All Results'!F4:F1000)"),
        ("Avg Iterations (GPT-4)",   "=AVERAGEIF('All Results'!D4:D1000,\"GPT-4\",'All Results'!F4:F1000)"),
        ("Avg Iterations (Gemini)",  "=AVERAGEIF('All Results'!D4:D1000,\"Gemini\",'All Results'!F4:F1000)"),
        ("Total AWS Cost Saved",     "=SUM('All Results'!K4:K1000)"),
        ("SUCCESS count",            "=COUNTIF('All Results'!I4:I1000,\"SUCCESS\")"),
        ("ESCALATED count",          "=COUNTIF('All Results'!I4:I1000,\"ESCALATED\")"),
        ("FAILED count",             "=COUNTIF('All Results'!I4:I1000,\"FAILED\")"),
    ]
    for i, (label, formula) in enumerate(summary, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True, size=9)
        ws2.cell(row=i, column=1).fill = PatternFill("solid", fgColor="E8EEF4")
        ws2.cell(row=i, column=2, value=formula)
        ws2.row_dimensions[i].height = 18
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 20

    wb.save(EXCEL_FILE)
    print(f"Excel created: {EXCEL_FILE}")
    return wb

def add_result(record):
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
    else:
        wb = create_excel()

    ws = wb["All Results"]
    next_row = max(ws.max_row + 1, 4)

    thin = Side(style="thin", color="EEEEEE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    values = [
        record["experiment_id"], record["timestamp"],
        record["template_type"], record["model"],
        record["first_pass"], record["iterations"],
        record["errors_found"], record["error_types"],
        record["final_status"], record["time_seconds"],
        record["aws_cost_saved"], record["notes"],
    ]

    row_color = COLOR_ROW_ALT if next_row % 2 == 0 else "FFFFFF"

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.font = Font(size=8)

        if col == 4:
            if val == "Claude":
                cell.fill = PatternFill("solid", fgColor="EEEDFE")
                cell.font = Font(size=8, bold=True, color=COLOR_CLAUDE)
            elif val == "GPT-4":
                cell.fill = PatternFill("solid", fgColor="E6F1FB")
                cell.font = Font(size=8, bold=True, color=COLOR_GPT4)
            elif val == "Gemini":
                cell.fill = PatternFill("solid", fgColor="E1F5EE")
                cell.font = Font(size=8, bold=True, color=COLOR_GEMINI)
        elif col == 5:
            if val == "YES":
                cell.fill = PatternFill("solid", fgColor=COLOR_PASS)
                cell.font = Font(size=8, bold=True, color="3B6D11")
            else:
                cell.fill = PatternFill("solid", fgColor=COLOR_FAIL)
                cell.font = Font(size=8, bold=True, color="A32D2D")
        elif col == 9:
            if val == "SUCCESS":
                cell.fill = PatternFill("solid", fgColor=COLOR_PASS)
                cell.font = Font(size=8, bold=True, color="3B6D11")
            elif val == "FAILED":
                cell.fill = PatternFill("solid", fgColor=COLOR_FAIL)
                cell.font = Font(size=8, bold=True, color="A32D2D")
            elif val == "ESCALATED":
                cell.fill = PatternFill("solid", fgColor=COLOR_WARN)
                cell.font = Font(size=8, bold=True, color="854F0B")
        else:
            cell.fill = PatternFill("solid", fgColor=row_color)

    wb.save(EXCEL_FILE)

    file_exists = CSV_FILE.exists()
    with open(CSV_FILE, "a", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=record.keys())
        if not file_exists:
            writer.writeheader()
        writer.writerow(record)

    print(f"Result saved — Exp {record['experiment_id']} | {record['model']} | {record['final_status']}")

if __name__ == "__main__":
    print("=" * 55)
    print("  AI-CFN Results Tracker")
    print("=" * 55)
    create_excel()
    print("Excel ready — run pipeline.py to add results!")
